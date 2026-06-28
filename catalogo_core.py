## catalogo_core.py
# =========================================
# Funciones núcleo compartidas para limpieza
# y análisis de catálogos en ERP
# =========================================

from typing import Dict, List

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# =========================================
# Estilos base
# =========================================
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(vertical="center", wrap_text=True)


# =========================================
# Utilidades generales
# =========================================
def _norm(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _count_missing(df: pd.DataFrame, col: str) -> int:
    if col not in df.columns:
        return len(df)
    s = df[col].astype(str).replace("nan", "").replace("None", "").str.strip()
    return int((s == "").sum())


def _count_duplicates(df: pd.DataFrame, col: str) -> int:
    if col not in df.columns:
        return 0
    s = df[col].astype(str).replace("nan", "").replace("None", "").str.strip()
    s = s[s != ""]
    return int(s.duplicated().sum())


def _resolve_col(df: pd.DataFrame, candidates: List[str]) -> str:
    mapping = {c.lower().replace(" ", "").replace("_", ""): c for c in df.columns}
    for cand in candidates:
        key = cand.lower().replace(" ", "").replace("_", "")
        if key in mapping:
            return mapping[key]
    return ""


def _apply_excel_style(output_path: str, df: pd.DataFrame, required_cols: List[str]) -> None:
    wb = load_workbook(output_path)
    ws = wb.active

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for row_idx in range(2, len(df) + 2):
        for col_name in required_cols:
            if col_name in df.columns:
                value = df.iloc[row_idx - 2][col_name]
                if pd.isna(value) or str(value).strip() == "":
                    col_idx = list(df.columns).index(col_name) + 1
                    ws.cell(row=row_idx, column=col_idx).fill = YELLOW_FILL

    # Código corregido para el ajuste de ancho de columnas:
    for i, col in enumerate(df.columns, start=1):
        max_len = len(str(col))
    for v in df.iloc[:, i - 1].head(200):
        # Convertimos explícitamente a string y manejamos nulos
        val_str = str(v) if not pd.isna(v) else ""
        max_len = max(max_len, len(val_str))

    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(output_path)


# =========================================
# Análisis catálogo productos
# =========================================
def analizar_catalogo(df: pd.DataFrame) -> Dict[str, int]:
    return {
        "filas_originales": int(len(df)),
        "sin_sku": _count_missing(df, "SKU"),
        "sin_nombre": _count_missing(df, "Nombre"),
        "sin_categoria": _count_missing(df, "Categoria"),
        "sin_precio": _count_missing(df, "Precio"),
        "sin_costo": _count_missing(df, "Costo1"),
        "sin_proveedor": _count_missing(df, "Proveedor1"),
        "sin_estado": _count_missing(df, "Estado"),
        "skus_duplicados": _count_duplicates(df, "SKU"),
    }


# =========================================
# Análisis materias primas
# =========================================
def analizar_materias_primas(df: pd.DataFrame) -> Dict[str, int]:
    return {
        "filas_originales": int(len(df)),
        "sin_nombre": _count_missing(df, "Nombre"),
        "sin_unidadmedida": _count_missing(df, "UnidadMedida"),
        "sin_unidadcompra": _count_missing(df, "UnidadCompra"),
        "sin_contenido": _count_missing(df, "Contenido"),
        "sin_proveedor": _count_missing(df, "Proveedor1"),
        "sin_costo": _count_missing(df, "Costo1"),
        "sin_stock": _count_missing(df, "Stock"),
        "sin_descripcion": _count_missing(df, "Descripcion"),
        "duplicados": _count_duplicates(df, "Nombre"),
    }


# =========================================
# Limpieza de productos
# =========================================
def limpiar_catalogo_productos(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    required = ["SKU", "Nombre", "Categoria", "Modelo", "Precio", "Costo1", "Proveedor1", "Estado"]
    for col in required:
        if col not in df.columns:
            df[col] = ""

    for col in required:
        df[col] = df[col].apply(_norm)

    df["Nombre"] = df["Nombre"].replace("", "SIN NOMBRE")
    df["Categoria"] = df["Categoria"].replace("", "SIN CATEGORIA")
    df["Modelo"] = df["Modelo"].replace("", "SIN MODELO")
    df["Estado"] = df["Estado"].replace("", "ACTIVO")

    return df


# =========================================
# Limpieza de materias primas
# =========================================
def limpiar_materias_primas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    c_nombre = _resolve_col(df, ["Nombre"])
    c_um = _resolve_col(df, ["UnidadMedida", "Unidad de medida", "Unidad medida", "UnidadUso"])
    c_uc = _resolve_col(df, ["UnidadCompra", "Unidad de compra", "Unidad compra", "UnidadEmpaque"])
    c_cont = _resolve_col(df, ["Contenido"])
    c_prov = _resolve_col(df, ["Proveedor1", "Proveedor 1", "Proveedor"])
    c_costo = _resolve_col(df, ["Costo1", "Costo 1", "Costo"])
    c_stock = _resolve_col(df, ["Stock"])
    c_desc = _resolve_col(df, ["Descripcion"])

    required = ["Nombre", "UnidadMedida", "UnidadCompra", "Contenido", "Proveedor1", "Costo1", "Stock", "DescripcionCompra"]
    for col in required:
        if col not in df.columns:
            df[col] = ""

    if c_nombre:
        df["Nombre"] = df[c_nombre].apply(_norm)
    if c_um:
        df["UnidadMedida"] = df[c_um].apply(_norm)
    if c_uc:
        df["UnidadCompra"] = df[c_uc].apply(_norm)
    if c_cont:
        df["Contenido"] = df[c_cont].apply(_norm)
    if c_prov:
        df["Proveedor1"] = df[c_prov].apply(_norm)
    if c_costo:
        df["Costo1"] = df[c_costo].apply(_norm)
    if c_stock:
        df["Stock"] = df[c_stock].apply(_norm)
    if c_desc:
        df["Descripcion"] = df[c_desc].apply(_norm)

    df["Nombre"] = df["Nombre"].replace("", "SIN NOMBRE")
    df["Proveedor1"] = df["Proveedor1"].replace("", "SIN PROVEEDOR")
    df["Costo1"] = df["Costo1"].replace("", "0")
    df["Stock"] = df["Stock"].replace("", "0")

    for idx in df.index:
        um = _norm(df.at[idx, "UnidadMedida"])
        uc = _norm(df.at[idx, "UnidadCompra"])
        cont = _norm(df.at[idx, "Contenido"])

        if um == "":
            continue

        if uc == "" and cont == "":
            df.at[idx, "UnidadCompra"] = um
            df.at[idx, "Contenido"] = "1"
            uc = um
            cont = "1"

        uc = _norm(df.at[idx, "UnidadCompra"])
        cont = _norm(df.at[idx, "Contenido"])

        if uc == "" or cont == "" or um == "":
            df.at[idx, "DescripcionCompra"] = ""
            continue

        if uc.lower() == um.lower():
            df.at[idx, "DescripcionCompra"] = ""
            continue

        df.at[idx, "DescripcionCompra"] = f"{uc} con {cont} {um}"

    final_cols = [
        "Nombre",
        "DescripcionCompra",
        "UnidadMedida",
        "UnidadCompra",
        "Contenido",
        "Proveedor1",
        "Costo1",
        "Stock",
    ]

    if "Descripcion" in df.columns:
        final_cols.insert(1, "Descripcion")

    rest = [c for c in df.columns if c not in final_cols]
    df = df[final_cols + rest]

    return df


# =========================================
# Exportación a Excel
# =========================================
def exportar_excel_con_estilo(df: pd.DataFrame, output_path: str, required_cols: List[str]) -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Catalogo")

    _apply_excel_style(output_path, df, required_cols)
    return output_path